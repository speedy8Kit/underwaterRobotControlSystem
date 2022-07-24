import socket

from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QListWidgetItem, QMessageBox
from openpyxl import Workbook

from listings.manager import manager
from listings.point import Point, list_point
from listings.ui.new_point_ui import PointWidget
from lyb import address, commandList


class SMUI_BtClickEvent(object):
    __number_point = 0
    __current_row = None
    main_window: QtWidgets.QMainWindow = None
    set_mission_window: QtWidgets.QMainWindow = None

    @staticmethod
    def addPoint(x=0, y=0, move_mode=2, height_mode=1):
        point = Point(x, y, move_mode, height_mode)
        point_widget = PointWidget(point, SMUI_BtClickEvent.__number_point)
        SMUI_BtClickEvent.__number_point += 1
        item = QListWidgetItem(PointWidget.list_point_widget)
        item.setSizeHint(point_widget.size())
        PointWidget.list_point_widget.setItemWidget(item, point_widget)

    @staticmethod
    def removePoint():
        if SMUI_BtClickEvent.__current_row is not None:
            try:
                list_point.remove(list_point[SMUI_BtClickEvent.__current_row])
                SMUI_BtClickEvent.__number_point -= 1
                PointWidget.list_point_widget.setCurrentRow(SMUI_BtClickEvent.__current_row)
            except:
                print("error deleted")
        SMUI_BtClickEvent.updateListPointWidgets()

    @staticmethod
    def updateListPointWidgets():
        PointWidget.list_point_widget.clear()
        number_point = 0
        for point in list_point:
            point_widget = PointWidget(point, number_point)
            number_point += 1
            item = QListWidgetItem(PointWidget.list_point_widget)
            item.setSizeHint(point_widget.size())
            PointWidget.list_point_widget.setItemWidget(item, point_widget)

    @staticmethod
    def rowChanged():
        SMUI_BtClickEvent.__current_row = PointWidget.list_point_widget.currentRow()

    @staticmethod
    def scan():
        list_point.clear()
        SMUI_BtClickEvent.addPoint()
        x = Point.x_min + 4
        y = Point.y_min
        while x < Point.x_max - 2:
            if y == Point.y_min:
                SMUI_BtClickEvent.addPoint(x, y + 4, 1, 1)
                y = Point.y_max
                SMUI_BtClickEvent.addPoint(x, y - 4, 1, 1)
            else:
                SMUI_BtClickEvent.addPoint(x, y - 4, 1, 1)
                y = Point.y_min
                SMUI_BtClickEvent.addPoint(x, y + 4, 1, 1)
            x += 8

    @staticmethod
    def cancel():
        list_point.clear()
        SMUI_BtClickEvent.__number_point = 0
        SMUI_BtClickEvent.addPoint()
        SMUI_BtClickEvent.updateListPointWidgets()
        SMUI_BtClickEvent.main_window.show()
        SMUI_BtClickEvent.set_mission_window.close()

    @staticmethod
    def save():
        if len(list_point) > 2:
            wb = Workbook()
            ws = wb.active

            i = 1
            ws[f'A{i}'] = "x"
            ws[f'B{i}'] = "y"
            ws[f'C{i}'] = "h"
            ws[f'D{i}'] = "alpha"
            ws[f'E{i}'] = "height_mode"
            ws[f'F{i}'] = "move_mode"

            for p in list_point:
                manager.sock.send_com(commandList.SRV_ADD_POINT(p.x, p.y, p.h, p.v, p.move_mode, p.height_mode, i - 1),
                                      address.addr_CS)
                i += 1
                ws[f'A{i}'] = p.x
                ws[f'B{i}'] = p.y
                ws[f'C{i}'] = p.h
                ws[f'D{i}'] = p.alpha
                ws[f'E{i}'] = p.height_mode
                ws[f'F{i}'] = p.move_mode

            ws2 = wb.create_sheet("holst_param")
            ws2[f'A1'] = "x min"
            ws2[f'B1'] = "x max"
            ws2[f'C1'] = "y min"
            ws2[f'D1'] = "x max"

            ws2[f'A2'] = Point.x_min
            ws2[f'B2'] = Point.x_max
            ws2[f'C2'] = Point.y_min
            ws2[f'D2'] = Point.y_max

            # Save the file
            wb.save("sample.xlsx")
            SMUI_BtClickEvent.main_window.ui.label_mission.setText("sample.xlsx")

            SMUI_BtClickEvent.main_window.show()
            SMUI_BtClickEvent.set_mission_window.close()
            manager.sock.send_com(commandList.SRV_GENERATE_PATH(len(list_point)), address.addr_CS)
            try:
                com, addr = manager.sock.recv_com()
                com.run(addr, manager.sock)
            except socket.timeout:
                print("fail")
        else:
            SMUI_BtClickEvent.error()

    @staticmethod
    def error():
        err = QMessageBox()
        err.setWindowTitle("БОЛЬШАЯ ОШИБКА")
        err.setText("кнопка пока не добавлена")
        err.setIcon(QMessageBox.Information)
        err.setStandardButtons(QMessageBox.Cancel | QMessageBox.Ok)
        err.exec()
