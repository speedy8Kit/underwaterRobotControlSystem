from PyQt5 import QtWidgets
from openpyxl import Workbook

import sim
from listings.manager import manager


class SSUI_BtClickEvent(object):
    main_window: QtWidgets.QMainWindow = None
    start_simulation_window: QtWidgets.QMainWindow = None

    @staticmethod
    def stopSimulation():
        sim.simxStopSimulation(manager.clientID, sim.simx_opmode_oneshot_wait)
        manager.worked = False
        SSUI_BtClickEvent.main_window.ui.label_connect.setText("repeat")
        SSUI_BtClickEvent.main_window.show()
        SSUI_BtClickEvent.start_simulation_window.close()

    @staticmethod
    def save():
        wb = Workbook()
        ws = wb.create_sheet("object")

        i = 1
        ws[f'A{i}'] = "x"
        ws[f'B{i}'] = "y"
        ws[f'C{i}'] = "z"
        for obj in manager.list_red_objects:
            i += 1
            ws[f'A{i}'] = obj.pos[0]
            ws[f'B{i}'] = obj.pos[1]
            ws[f'C{i}'] = obj.pos[2]

        wb.save("obj.xlsx")